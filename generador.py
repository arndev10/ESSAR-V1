import os
import pandas as pd
from openpyxl import load_workbook

MAPEO_CELDAS = {
    "A": "L11",
    "B": "C26",
    "C": "D26",
    "D": "F26",
    "E": "G33",
    "F": "H50",
    "G": "F21",
    "H": "L21"
}


def generar(ruta_datos, plantilla_base, carpeta_salida, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    if not os.path.isdir(ruta_datos):
        raise FileNotFoundError(f"No existe la carpeta de insumos: {ruta_datos}")
    if not os.path.isfile(plantilla_base):
        raise FileNotFoundError(f"No existe la plantilla: {plantilla_base}")

    os.makedirs(carpeta_salida, exist_ok=True)
    archivos_xlsx = [f for f in os.listdir(ruta_datos) if f.endswith(".xlsx")]
    total_archivos = 0

    for archivo in archivos_xlsx:
        ruta_archivo = os.path.join(ruta_datos, archivo)
        log(f"Procesando: {archivo}")

        df = pd.read_excel(ruta_archivo, skiprows=1, header=None)
        carpeta_archivo = os.path.join(carpeta_salida, os.path.splitext(archivo)[0])
        os.makedirs(carpeta_archivo, exist_ok=True)

        for i, fila in df.iterrows():
            wb = load_workbook(plantilla_base)
            ws = wb.active
            for idx, col in enumerate("ABCDEFGH"):
                valor = fila[idx]
                celda = MAPEO_CELDAS[col]
                ws[celda] = valor
            nombre_archivo = f"ESAR__{fila[0]}__{fila[1]}__{fila[2]}.xlsx"
            ruta_salida = os.path.join(carpeta_archivo, nombre_archivo)
            wb.save(ruta_salida)
            total_archivos += 1

    log(f"Listo. Se generaron {total_archivos} archivos.")
    return total_archivos
